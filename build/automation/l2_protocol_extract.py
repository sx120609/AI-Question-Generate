from __future__ import annotations

import argparse
import hashlib
import json
import random
import sys
from pathlib import Path

from openpyxl import load_workbook


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_text(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return path.read_text(encoding=encoding).strip()
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Unable to decode text input: {path}")


def resolve_input(root: Path, preferred: str, fallback: str | None = None):
    preferred_path = root / preferred
    if preferred_path.is_file():
        return preferred_path, False
    if fallback:
        fallback_path = root / fallback
        if fallback_path.is_file():
            return fallback_path, True
    return preferred_path, False


def load_reference_rows(path: Path, question_prefix: str, attachment_header: str):
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = []
    sheet_reports = []
    for sheet in workbook.worksheets:
        # The supplied workbook declares dimension A1 although it contains more
        # than a thousand rows. Without this reset, read-only loaders silently
        # expose an empty workbook.
        sheet.reset_dimensions()
        iterator = sheet.iter_rows(values_only=True)
        header_values = next(iterator, ())
        headers = [str(value).strip() if value is not None else "" for value in header_values]
        question_index = next((index for index, value in enumerate(headers) if value.startswith(question_prefix)), None)
        attachment_index = next((index for index, value in enumerate(headers) if value == attachment_header), None)
        if question_index is None or attachment_index is None:
            sheet_reports.append({"sheet": sheet.title, "status": "SKIPPED", "reason": "required headers not found"})
            continue
        valid = 0
        incomplete = 0
        for row_number, values in enumerate(iterator, 2):
            question = str(values[question_index]).strip() if question_index < len(values) and values[question_index] is not None else ""
            attachment_summary = str(values[attachment_index]).strip() if attachment_index < len(values) and values[attachment_index] is not None else ""
            if question and attachment_summary:
                valid += 1
                rows.append({
                    "sheet": sheet.title,
                    "row": row_number,
                    "question": question,
                    "attachmentSummary": attachment_summary,
                    "questionHash": hashlib.sha256(question.encode("utf-8")).hexdigest(),
                    "attachmentSummaryHash": hashlib.sha256(attachment_summary.encode("utf-8")).hexdigest(),
                })
            elif question or attachment_summary:
                incomplete += 1
        sheet_reports.append({
            "sheet": sheet.title,
            "status": "READY",
            "questionColumn": question_index + 1,
            "attachmentSummaryColumn": attachment_index + 1,
            "validRows": valid,
            "incompleteRows": incomplete,
        })
    return rows, sheet_reports


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", required=True)
    parser.add_argument("--run-id", required=True)
    parser.add_argument("--count", type=int, required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--seed", default="")
    args = parser.parse_args()

    root = Path(args.root).resolve()
    out_path = Path(args.out).resolve()
    config_path = root / "config" / "l2_production_protocol.json"
    config = json.loads(config_path.read_text(encoding="utf-8"))
    configured = config["inputs"]
    requirements_path, requirements_fallback = resolve_input(
        root,
        configured["requirementsPreferred"],
        configured.get("requirementsNormalizedFallback"),
    )
    paths = {
        "requirements": requirements_path,
        "referenceWorkbook": root / configured["referenceWorkbook"],
        "firstQaPrompt": root / configured["firstQaPrompt"],
        "secondQaPrompt": root / configured["secondQaPrompt"],
    }
    missing = [str(path) for path in paths.values() if not path.is_file()]
    packet = {
        "schemaVersion": 1,
        "kind": "l2-production-input-packet",
        "protocolId": config["protocolId"],
        "runId": args.run_id,
        "questionCount": args.count,
        "status": "BLOCKED" if missing else "READY",
        "blockers": [{"rule": "required-input-missing", "path": value} for value in missing],
        "warnings": ([{
            "rule": "normalized-requirements-fallback",
            "message": "Original DOCX is absent; using the repository-normalized Markdown extraction.",
            "path": str(requirements_path),
        }] if requirements_fallback and not missing else []),
    }
    if missing:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(packet, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(json.dumps(packet, ensure_ascii=False))
        return 2

    reference_rows, sheet_reports = load_reference_rows(
        paths["referenceWorkbook"],
        config["referenceColumns"]["questionHeaderPrefix"],
        config["referenceColumns"]["attachmentSummaryHeader"],
    )
    if len(reference_rows) < args.count:
        packet["status"] = "BLOCKED"
        packet["blockers"].append({
            "rule": "insufficient-reference-rows",
            "available": len(reference_rows),
            "required": args.count,
        })
    seed_text = args.seed or f"{args.run_id}:{sha256_file(paths['referenceWorkbook'])}"
    rng = random.Random(int(hashlib.sha256(seed_text.encode("utf-8")).hexdigest(), 16))
    selected = rng.sample(reference_rows, min(args.count, len(reference_rows)))
    packet["inputs"] = {
        "requirements": {
            "path": str(paths["requirements"]),
            "sha256": sha256_file(paths["requirements"]),
            "normalizedFallback": requirements_fallback,
            "sourceUrl": configured.get("requirementsRemoteSource", ""),
            "text": read_text(paths["requirements"]),
        },
        "referenceWorkbook": {
            "path": str(paths["referenceWorkbook"]),
            "sha256": sha256_file(paths["referenceWorkbook"]),
            "selectedColumnsOnly": [
                config["referenceColumns"]["questionAlias"],
                config["referenceColumns"]["attachmentSummaryHeader"],
            ],
            "eligibleRows": len(reference_rows),
            "sheets": sheet_reports,
            "samples": [
                {"questionIndex": index + 1, **sample}
                for index, sample in enumerate(selected)
            ],
        },
        "firstQaPrompt": {
            "path": str(paths["firstQaPrompt"]),
            "sha256": sha256_file(paths["firstQaPrompt"]),
            "text": read_text(paths["firstQaPrompt"]),
        },
        "secondQaPrompt": {
            "path": str(paths["secondQaPrompt"]),
            "sha256": sha256_file(paths["secondQaPrompt"]),
            "text": read_text(paths["secondQaPrompt"]),
        },
    }
    packet["requiredTraceSections"] = [
        "referenceLocation",
        "referenceQuestionStructure",
        "referenceAttachmentStructure",
        "newQuestionStructureMapping",
        "newAttachmentSupport",
        "firstQaFullResult",
        "firstQaRepairs",
        "secondQaFullResult",
        "finalRecord",
    ]
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(packet, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "status": packet["status"],
        "runId": args.run_id,
        "eligibleRows": len(reference_rows),
        "sampledRows": [{"sheet": row["sheet"], "row": row["row"]} for row in selected],
        "warnings": packet["warnings"],
    }, ensure_ascii=False))
    return 0 if packet["status"] == "READY" else 2


if __name__ == "__main__":
    sys.exit(main())
