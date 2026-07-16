"""
数据加载器 - 加载规则文档、Prompt、GoodCase/BadCase
"""
import re
import json
import yaml
from pathlib import Path
from typing import List, Dict, Optional
import openpyxl


# ========== 规则加载 ==========

def load_rules_text(filepath: str) -> str:
    """加载规则文档原始文本"""
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"规则文档不存在: {filepath}")
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def parse_rules_structured(filepath: str) -> List[Dict]:
    """将规则文档解析为结构化规则列表"""
    text = load_rules_text(filepath)
    rules = []

    # 提取规则编号和描述
    rule_pattern = re.compile(
        r'(\d+)\.\s*(.+?)(?:（(严重|中等|轻微).*?）)?\n'
        r'(.*?)(?=\n\d+\.\s|\n[A-Z]+\.\s|\n【|$)',
        re.DOTALL
    )

    for match in rule_pattern.finditer(text):
        rule_id = f"R{match.group(1)}"
        title = match.group(2).strip()
        severity = match.group(3) or "中等"
        detail = match.group(4).strip()

        # 提取"必须判错"和"不要误判"
        must_flag = "必须判错" in detail
        dont_misjudge = "不要误判" in detail

        rules.append({
            "id": rule_id,
            "title": title,
            "severity": severity,
            "detail": detail[:500],  # 截断过长内容
            "must_flag": must_flag,
            "dont_misjudge": dont_misjudge,
        })

    # 补充规则 A, B, C
    supplement_pattern = re.compile(
        r'([A-C]+)\.\s*(.+?)(?:（(严重|中等|轻微).*?）)?\n'
        r'(.*?)(?=\n[A-C]+\.\s|\n\d+\.\s|$)',
        re.DOTALL
    )
    for match in supplement_pattern.finditer(text):
        rule_id = f"R{match.group(1)}"
        title = match.group(2).strip()
        severity = match.group(3) or "中等"
        detail = match.group(4).strip()
        rules.append({
            "id": rule_id,
            "title": title,
            "severity": severity,
            "detail": detail[:500],
            "must_flag": "必须判错" in detail,
            "dont_misjudge": "不要误判" in detail,
        })

    return rules


# ========== Prompt 加载 ==========

def load_qa_prompt(filepath: str) -> str:
    """加载质检 Prompt 原文"""
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Prompt文件不存在: {filepath}")
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


# ========== Case 加载（从 Excel 反馈表）==========

def load_cases_from_excel(filepath: str) -> Dict[str, List[Dict]]:
    """
    从反馈表.xlsx 加载 GoodCase 和 BadCase

    GoodCase = 未被打回/无质检意见的题目（质检状态为空或"通过"）
    BadCase = 有质检意见的题目

    Returns:
        {"good": [...], "bad": [...]}
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"反馈表不存在: {filepath}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    good_cases = []
    bad_cases = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        uid = str(row[0] or "").strip()
        title = str(row[1] or "").strip()
        task_type = str(row[2] or "").strip()
        category_l1 = str(row[3] or "").strip()
        category_l2 = str(row[4] or "").strip()
        category_l3 = str(row[5] or "").strip()
        task_summary = str(row[6] or "").strip()
        attachment_desc = str(row[9] or "").strip()
        output_format = str(row[12] or "").strip()
        expert_name = str(row[16] or "").strip()
        review_status = str(row[18] or "").strip() if row[18] else ""
        review_opinion = str(row[19] or "").strip() if row[19] else ""

        if not uid or not title:
            continue

        case = {
            "uid": uid,
            "title": title[:2000],  # 截断超长题目
            "task_type": task_type,
            "category": f"{category_l1} > {category_l2} > {category_l3}",
            "task_summary": task_summary,
            "attachment_desc": attachment_desc[:500],
            "output_format": output_format,
            "expert_name": expert_name,
        }

        if review_opinion and review_opinion.strip():
            # 有质检意见 = BadCase
            case["issues"] = review_opinion
            bad_cases.append(case)
        else:
            # 无质检意见 = GoodCase（通过了质检）
            good_cases.append(case)

    print(f"📊 加载完成: {len(good_cases)} 条 GoodCase, {len(bad_cases)} 条 BadCase")
    return {"good": good_cases, "bad": bad_cases}


# ========== 测试集提取 ==========

def load_test_set(filepath: str, limit: int = 0) -> List[Dict]:
    """
    加载待质检的题目（质检状态为空的题目，或全部题目）
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"反馈表不存在: {filepath}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    test_items = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        uid = str(row[0] or "").strip()
        title = str(row[1] or "").strip()
        task_type = str(row[2] or "").strip()
        review_status = str(row[18] or "").strip() if row[18] else ""
        review_opinion = str(row[19] or "").strip() if row[19] else ""

        if not uid or not title:
            continue

        # 只取待质检的（没有质检意见的）
        if not review_opinion:
            test_items.append({
                "uid": uid,
                "title": title[:2000],
                "task_type": task_type,
                "expected_review": review_status,
            })

        if limit and len(test_items) >= limit:
            break

    return test_items


if __name__ == "__main__":
    # 测试
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    from config import RULES_PATH, PROMPT_PATH, CASES_PATH

    rules = parse_rules_structured(RULES_PATH)
    print(f"\n规则数: {len(rules)}")
    for r in rules:
        print(f"  {r['id']}: {r['title'][:60]} [{r['severity']}]")

    cases = load_cases_from_excel(CASES_PATH)
    print(f"\nGoodCase 示例:")
    if cases["good"]:
        print(f"  UID: {cases['good'][0]['uid']}")
        print(f"  题目: {cases['good'][0]['title'][:100]}...")
    print(f"\nBadCase 示例:")
    if cases["bad"]:
        print(f"  UID: {cases['bad'][0]['uid']}")
        print(f"  问题: {cases['bad'][0]['issues'][:200]}...")
